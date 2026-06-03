#!/usr/bin/env python3
"""
Update ww2_draft_updated.xlsx to copy Ancestry URLs from familysearch_citation
to ancestry_url column when ancestry_url is blank.
"""

import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from loguru import logger


def extract_ancestry_url(text: str) -> str | None:
    """Extract Ancestry URL from citation text."""
    if not text:
        return None

    # Look for ancestrylibrary.com URL
    match = re.search(r'https://www\.ancestrylibrary\.com/[^\s\)]+', text)
    if match:
        # Clean up trailing punctuation
        url = match.group(0).rstrip('.,;)\'"')
        return url

    return None


def is_blank(value) -> bool:
    """Check if a cell value is blank/empty."""
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def main():
    """Main update function."""
    xlsx_path = Path(__file__).parent.parent / "ww2_draft_updated.xlsx"

    if not xlsx_path.exists():
        logger.error(f"File not found: {xlsx_path}")
        return 1

    # Create backup
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = xlsx_path.with_name(f"ww2_draft_updated_backup_{timestamp}.xlsx")
    shutil.copy2(xlsx_path, backup_path)
    logger.info(f"✅ Created backup: {backup_path.name}")

    # Load workbook
    logger.info(f"📖 Loading {xlsx_path.name}...")
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb.active

    # Find column indexes
    headers = [cell.value for cell in sheet[1]]
    try:
        fs_citation_col = headers.index("familysearch_citation") + 1  # 1-indexed
        ancestry_url_col = headers.index("ancestry_url") + 1
    except ValueError as e:
        logger.error(f"Required column not found: {e}")
        wb.close()
        return 1

    logger.info(f"   familysearch_citation column: {fs_citation_col}")
    logger.info(f"   ancestry_url column: {ancestry_url_col}")

    # Process rows
    updated_count = 0
    skipped_count = 0
    total_rows = sheet.max_row - 1  # Exclude header

    logger.info(f"\n🔄 Processing {total_rows} data rows...")

    for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
        fs_citation = sheet.cell(row_idx, fs_citation_col).value
        ancestry_url = sheet.cell(row_idx, ancestry_url_col).value

        # Check if ancestry_url is blank
        if not is_blank(ancestry_url):
            continue

        # Check if familysearch_citation contains Ancestry URL
        if fs_citation and "https://www.ancestrylibrary.com" in str(fs_citation):
            extracted_url = extract_ancestry_url(str(fs_citation))

            if extracted_url:
                # Copy URL to ancestry_url column
                sheet.cell(row_idx, ancestry_url_col).value = extracted_url
                updated_count += 1

                # Log first few updates for verification
                if updated_count <= 5:
                    logger.info(f"   Row {row_idx}: Copied URL")
                    logger.info(f"      → {extracted_url[:80]}...")
            else:
                skipped_count += 1
                logger.warning(f"   Row {row_idx}: Contains ancestrylibrary.com but URL extraction failed")

    # Save changes
    logger.info(f"\n💾 Saving changes to {xlsx_path.name}...")
    wb.save(xlsx_path)
    wb.close()

    # Report results
    logger.info("\n" + "="*60)
    logger.info("✅ Update Complete")
    logger.info("="*60)
    logger.info(f"   Total rows processed: {total_rows}")
    logger.info(f"   Rows updated: {updated_count}")
    logger.info(f"   Rows skipped (extraction failed): {skipped_count}")
    logger.info(f"   Rows unchanged: {total_rows - updated_count - skipped_count}")
    logger.info(f"\n   Backup saved: {backup_path.name}")
    logger.info(f"   Updated file: {xlsx_path.name}")

    return 0


if __name__ == "__main__":
    exit(main())
