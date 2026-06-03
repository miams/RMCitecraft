#!/usr/bin/env python3
"""
Extract Census Source Information from RootsMagic to XLSX.

This script extracts U.S. Federal Census source records from a RootsMagic
genealogy database and exports them to an Excel spreadsheet. It creates
separate worksheets for each census year (1910, 1920, 1930, 1940, 1950).

For each source, the script extracts:
  - Source details: name, footnote, short footnote, bibliography
  - Person info: RIN, surname, given name (first person linked via census citation)
  - Citation details: name, page, research note, webtags
  - Media info: filename(s), caption, description, date

Features:
  - Converts <i>...</i> italic tags to native Excel italic formatting
  - Converts RootsMagic date format to native Excel dates
  - Decodes HTML entities in citation text

Examples:
    # Extract using default database (data/Iiams.rmtree)
    uv run python scripts/extract_census_sources_to_xlsx.py

    # Specify custom database
    uv run python scripts/extract_census_sources_to_xlsx.py --db path/to/database.rmtree

    # Specify output filename
    uv run python scripts/extract_census_sources_to_xlsx.py --output census_report.xlsx

    # Extract only specific years
    uv run python scripts/extract_census_sources_to_xlsx.py --years 1940 1950
"""

import argparse
import re
import sqlite3
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: uv pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Census years to extract
CENSUS_YEARS = [1910, 1920, 1930, 1940, 1950]

# Census FactTypeID in RootsMagic
CENSUS_FACT_TYPE_ID = 18

# Regex pattern for RootsMagic date format: D.+YYYYMMDD..+00000000..
RM_DATE_PATTERN = re.compile(r'D\.\+(\d{4})(\d{2})(\d{2})\.\.\+\d{8}\.\.')


def parse_rm_date(rm_date: str) -> date | None:
    """Parse RootsMagic date string to Python date object.

    RootsMagic stores dates in format: D.+YYYYMMDD..+00000000..
    Example: D.+19400401..+00000000.. = April 1, 1940

    Args:
        rm_date: RootsMagic date string

    Returns:
        Python date object, or None if parsing fails or date is empty/invalid
    """
    if not rm_date or not isinstance(rm_date, str):
        return None

    match = RM_DATE_PATTERN.match(rm_date)
    if not match:
        return None

    try:
        year = int(match.group(1))
        month = int(match.group(2))
        day = int(match.group(3))

        # Handle partial dates (month=0 or day=0 means unknown)
        if month == 0 or day == 0:
            # Use January 1 for year-only dates
            month = max(1, month)
            day = max(1, day)

        return date(year, month, day)

    except (ValueError, IndexError):
        return None


@dataclass
class CensusSourceRecord:
    """Data class for a census source record."""
    # Source info
    source_id: int
    source_name: str
    footnote: str
    short_footnote: str
    bibliography: str
    source_text: str
    source_comment: str
    source_ref: str
    num_citations: int

    # Person info (first RIN linked via census citation)
    rin: int | None
    surname: str
    given_name: str

    # Citation info (linked to census event)
    citation_id: int | None
    citation_name: str
    page_number: str
    research_note: str
    detail_comment: str
    detail_ref: str
    citation_webtags: str
    citation_usage_count: int

    # Media info (first media item linked to source)
    media_filename: str
    num_media: int
    media_caption: str
    media_description: str
    media_date: date | None  # Converted from RootsMagic format to Python date
    media_ref: str
    media_scrapbook: str
    media_primary: str


def connect_database(db_path: Path) -> sqlite3.Connection:
    """Connect to RootsMagic database with ICU extension for RMNOCASE."""
    if not db_path.exists():
        raise FileNotFoundError(f"Database not found: {db_path}")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    # Try to load ICU extension for RMNOCASE collation
    script_dir = Path(__file__).parent.parent
    possible_paths = [
        script_dir / 'sqlite-extension/icu.dylib',
        Path('sqlite-extension/icu.dylib'),
        Path.cwd() / 'sqlite-extension/icu.dylib',
    ]

    for icu_path in possible_paths:
        if icu_path.exists():
            try:
                conn.enable_load_extension(True)
                conn.load_extension(str(icu_path))
                conn.execute(
                    "SELECT icu_load_collation("
                    "'en_US@colStrength=primary;caseLevel=off;normalization=on',"
                    "'RMNOCASE')"
                )
                conn.enable_load_extension(False)
                break
            except Exception as e:
                print(f"Warning: Could not load ICU extension from {icu_path}: {e}", file=sys.stderr)

    return conn


def extract_field_from_blob(fields_blob: bytes | str | None, field_name: str) -> str:
    """Extract a field value from the Fields BLOB XML structure.

    The BLOB contains XML like:
    <Root><Fields>
      <Field><Name>Footnote</Name><Value>...</Value></Field>
      ...
    </Fields></Root>
    """
    if not fields_blob:
        return ""

    try:
        fields_text = fields_blob.decode('utf-8', errors='ignore') if isinstance(fields_blob, bytes) else fields_blob
        # Remove BOM if present
        if fields_text.startswith('\ufeff'):
            fields_text = fields_text[1:]
        pattern = rf'<Name>{re.escape(field_name)}</Name>\s*<Value>(.*?)</Value>'
        match = re.search(pattern, fields_text, re.DOTALL)
        return match.group(1).strip() if match else ""
    except Exception:
        return ""


def get_census_sources(conn: sqlite3.Connection, year: int) -> list[CensusSourceRecord]:
    """Get all census sources for a specific year with associated information."""
    cursor = conn.cursor()

    # Get all sources for this census year
    cursor.execute('''
        SELECT
            s.SourceID,
            s.Name,
            s.Fields,
            s.ActualText,
            s.Comments,
            s.RefNumber,
            (SELECT COUNT(*) FROM CitationTable c WHERE c.SourceID = s.SourceID) as NumCitations
        FROM SourceTable s
        WHERE s.Name LIKE ?
        ORDER BY s.Name
    ''', (f'Fed Census: {year},%',))

    sources = cursor.fetchall()
    records = []

    for source in sources:
        source_id = source['SourceID']
        fields_blob = source['Fields']

        # Extract footnote, short footnote, bibliography from BLOB
        footnote = extract_field_from_blob(fields_blob, "Footnote")
        short_footnote = extract_field_from_blob(fields_blob, "ShortFootnote")
        bibliography = extract_field_from_blob(fields_blob, "Bibliography")

        # Get first citation linked to a census event and its person
        rin, surname, given_name = None, "", ""
        citation_id, citation_name, page_number = None, "", ""
        research_note, detail_comment, detail_ref = "", "", ""
        citation_webtags = ""
        citation_usage_count = 0

        # Find first citation linked to a census event (FactTypeID = 18)
        cursor.execute('''
            SELECT DISTINCT
                c.CitationID,
                c.CitationName,
                c.Fields as CitationFields,
                c.ActualText as ResearchNote,
                c.Comments as DetailComment,
                c.RefNumber as DetailRef,
                cl.OwnerID as EventID,
                e.OwnerID as PersonID
            FROM CitationTable c
            JOIN CitationLinkTable cl ON cl.CitationID = c.CitationID
            JOIN EventTable e ON e.EventID = cl.OwnerID AND cl.OwnerType = 2
            WHERE c.SourceID = ?
              AND e.EventType = ?
              AND e.OwnerType = 0
            ORDER BY c.CitationID
            LIMIT 1
        ''', (source_id, CENSUS_FACT_TYPE_ID))

        cit_row = cursor.fetchone()

        if cit_row:
            citation_id = cit_row['CitationID']
            citation_name = cit_row['CitationName'] or ""
            research_note = cit_row['ResearchNote'] or ""
            detail_comment = cit_row['DetailComment'] or ""
            detail_ref = cit_row['DetailRef'] or ""
            rin = cit_row['PersonID']

            # Extract "Page" field from citation fields blob
            cit_fields = cit_row['CitationFields']
            page_number = extract_field_from_blob(cit_fields, "Page")

            # Get person name
            if rin:
                cursor.execute('''
                    SELECT Surname, Given
                    FROM NameTable
                    WHERE OwnerID = ? AND IsPrimary = 1
                ''', (rin,))
                name_row = cursor.fetchone()
                if name_row:
                    surname = name_row['Surname'] or ""
                    given_name = name_row['Given'] or ""

            # Get citation webtags
            cursor.execute('''
                SELECT GROUP_CONCAT(Name || ': ' || URL, '; ') as Webtags
                FROM URLTable
                WHERE OwnerType = 4 AND OwnerID = ?
            ''', (citation_id,))
            webtag_row = cursor.fetchone()
            if webtag_row and webtag_row['Webtags']:
                citation_webtags = webtag_row['Webtags']

            # Get citation usage count
            cursor.execute('''
                SELECT COUNT(*) as cnt FROM CitationLinkTable WHERE CitationID = ?
            ''', (citation_id,))
            usage_row = cursor.fetchone()
            citation_usage_count = usage_row['cnt'] if usage_row else 0

        # Get media information (first media item linked to source)
        media_filename, num_media = "", 0
        media_caption, media_description, media_ref = "", "", ""
        media_date: date | None = None
        media_scrapbook, media_primary = "No", "No"

        # Count total media for source
        cursor.execute('''
            SELECT COUNT(*) as cnt
            FROM MediaLinkTable ml
            WHERE ml.OwnerType = 3 AND ml.OwnerID = ?
        ''', (source_id,))
        media_count_row = cursor.fetchone()
        num_media = media_count_row['cnt'] if media_count_row else 0

        # Get first media item details
        cursor.execute('''
            SELECT
                m.MediaFile,
                m.Caption,
                m.Description,
                m.Date,
                m.RefNumber,
                ml.Include1 as Scrapbook,
                ml.IsPrimary
            FROM MediaLinkTable ml
            JOIN MultimediaTable m ON m.MediaID = ml.MediaID
            WHERE ml.OwnerType = 3 AND ml.OwnerID = ?
            ORDER BY ml.SortOrder, ml.LinkID
            LIMIT 1
        ''', (source_id,))

        media_row = cursor.fetchone()
        if media_row:
            media_filename = media_row['MediaFile'] or ""
            media_caption = media_row['Caption'] or ""
            media_description = media_row['Description'] or ""
            media_date = parse_rm_date(media_row['Date'])  # Convert to native date
            media_ref = media_row['RefNumber'] or ""
            media_scrapbook = "Yes" if media_row['Scrapbook'] == 1 else "No"
            media_primary = "Yes" if media_row['IsPrimary'] == 1 else "No"

        # If multiple media, collect all filenames
        if num_media > 1:
            cursor.execute('''
                SELECT m.MediaFile
                FROM MediaLinkTable ml
                JOIN MultimediaTable m ON m.MediaID = ml.MediaID
                WHERE ml.OwnerType = 3 AND ml.OwnerID = ?
                ORDER BY ml.SortOrder, ml.LinkID
            ''', (source_id,))
            all_media = cursor.fetchall()
            media_filename = "; ".join(row['MediaFile'] for row in all_media if row['MediaFile'])

        record = CensusSourceRecord(
            source_id=source_id,
            source_name=source['Name'] or "",
            footnote=footnote,
            short_footnote=short_footnote,
            bibliography=bibliography,
            source_text=source['ActualText'] or "",
            source_comment=source['Comments'] or "",
            source_ref=source['RefNumber'] or "",
            num_citations=source['NumCitations'],
            rin=rin,
            surname=surname,
            given_name=given_name,
            citation_id=citation_id,
            citation_name=citation_name,
            page_number=page_number,
            research_note=research_note,
            detail_comment=detail_comment,
            detail_ref=detail_ref,
            citation_webtags=citation_webtags,
            citation_usage_count=citation_usage_count,
            media_filename=media_filename,
            num_media=num_media,
            media_caption=media_caption,
            media_description=media_description,
            media_date=media_date,
            media_ref=media_ref,
            media_scrapbook=media_scrapbook,
            media_primary=media_primary,
        )
        records.append(record)

    return records


def decode_html_entities(text: str) -> str:
    """Decode common HTML entities in text."""
    if not text:
        return text
    # Decode HTML entities
    text = text.replace('&lt;', '<')
    text = text.replace('&gt;', '>')
    text = text.replace('&quot;', '"')
    text = text.replace('&amp;', '&')  # Do this last to avoid double-decoding
    return text


def convert_italic_tags_to_rich_text(text: str) -> CellRichText | str:
    """Convert <i>text</i> tags to Excel rich text with italic formatting.

    Handles both raw tags (<i>) and HTML-encoded tags (&lt;i&gt;).
    Returns CellRichText if italic tags are found, otherwise returns plain string.
    """
    if not text:
        return text

    # First decode HTML entities so we have consistent tags
    decoded_text = decode_html_entities(text)

    if '<i>' not in decoded_text.lower():
        return decoded_text

    # Pattern to match italic tags
    pattern = r'<i>(.*?)</i>'

    # Find all parts (normal and italic)
    parts = []
    last_end = 0

    for match in re.finditer(pattern, decoded_text, re.IGNORECASE):
        # Add text before this italic section
        if match.start() > last_end:
            normal_text = decoded_text[last_end:match.start()]
            if normal_text:
                parts.append(TextBlock(InlineFont(), normal_text))

        # Add italic text
        italic_text = match.group(1)
        if italic_text:
            parts.append(TextBlock(InlineFont(i=True), italic_text))

        last_end = match.end()

    # Add any remaining text after last italic section
    if last_end < len(decoded_text):
        remaining_text = decoded_text[last_end:]
        if remaining_text:
            parts.append(TextBlock(InlineFont(), remaining_text))

    if parts:
        return CellRichText(parts)
    return decoded_text


def _get_text_content(text: str) -> str:
    """Get plain text content, stripping HTML/italic tags for length calculation."""
    if not text:
        return ""
    # Decode HTML entities and remove italic tags
    decoded = decode_html_entities(text)
    # Remove any remaining tags
    return re.sub(r'<[^>]+>', '', decoded)


def create_xlsx(records_by_year: dict[int, list[CensusSourceRecord]], output_path: Path) -> None:
    """Create XLSX file with separate sheets for each census year."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Column headers
    headers = [
        "RIN",
        "Surname",
        "Given Name",
        "Source Name",
        "Footnote",
        "Short Footnote",
        "Bibliography",
        "Number of Citations",
        "Source Text",
        "Source Comment",
        "Source Ref#",
        "Citation Name",
        "Page Number",
        "Research Note",
        "Detail Comment",
        "Detail Ref#",
        "Citation Webtags",
        "Number of times Citation used",
        "Media image filename",
        "Number of media images linked to Source",
        "Caption",
        "Description",
        "Date",
        "Ref#",
        "Scrapbook",
        "Primary",
    ]

    # Style definitions
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell_alignment = Alignment(vertical="top", wrap_text=False)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for year in CENSUS_YEARS:
        records = records_by_year.get(year, [])

        # Create worksheet
        ws = wb.create_sheet(title=str(year))

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, record in enumerate(records, start=2):
            row_data = [
                record.rin if record.rin else "",
                record.surname,
                record.given_name,
                record.source_name,
                convert_italic_tags_to_rich_text(record.footnote),
                record.short_footnote,
                convert_italic_tags_to_rich_text(record.bibliography),
                record.num_citations,
                record.source_text,
                record.source_comment,
                record.source_ref,
                record.citation_name,
                record.page_number,
                record.research_note,
                record.detail_comment,
                record.detail_ref,
                record.citation_webtags,
                record.citation_usage_count,
                record.media_filename,
                record.num_media,
                record.media_caption,
                record.media_description,
                record.media_date,
                record.media_ref,
                record.media_scrapbook,
                record.media_primary,
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                # Apply date format to Date column (column 23)
                if col == 23 and isinstance(value, date):
                    cell.number_format = 'YYYY-MM-DD'

        # Calculate column widths based on content
        # Start with header widths
        column_widths = [len(header) + 2 for header in headers]

        # Check all data rows to find max width for each column
        for record in records:
            row_data = [
                str(record.rin) if record.rin else "",
                record.surname,
                record.given_name,
                record.source_name,
                _get_text_content(record.footnote),
                record.short_footnote,
                _get_text_content(record.bibliography),
                str(record.num_citations),
                record.source_text,
                record.source_comment,
                record.source_ref,
                record.citation_name,
                record.page_number,
                record.research_note,
                record.detail_comment,
                record.detail_ref,
                record.citation_webtags,
                str(record.citation_usage_count),
                record.media_filename,
                str(record.num_media),
                record.media_caption,
                record.media_description,
                record.media_date.isoformat() if record.media_date else "",
                record.media_ref,
                record.media_scrapbook,
                record.media_primary,
            ]
            for col_idx, value in enumerate(row_data):
                if value:
                    # Add 2 characters padding for readability
                    column_widths[col_idx] = max(column_widths[col_idx], len(str(value)) + 2)

        # Apply column widths (cap at 100 to prevent extremely wide columns)
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = min(width, 100)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"

    # Save workbook
    wb.save(output_path)


def main() -> int:
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument(
        "--db",
        type=Path,
        default=Path("data/Iiams.rmtree"),
        help="Path to RootsMagic database (default: data/Iiams.rmtree)"
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output XLSX file path (default: census_sources_YYYYMMDD_HHMMSS.xlsx)"
    )

    parser.add_argument(
        "--years",
        type=int,
        nargs="+",
        default=CENSUS_YEARS,
        choices=CENSUS_YEARS,
        help=f"Census years to extract (default: {CENSUS_YEARS})"
    )

    args = parser.parse_args()

    # Generate default output filename if not specified
    if args.output is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = Path(f"census_sources_{timestamp}.xlsx")

    print(f"Connecting to database: {args.db}")

    try:
        conn = connect_database(args.db)
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    # Extract records for each year
    records_by_year = {}
    total_records = 0

    for year in args.years:
        print(f"Extracting {year} census sources...", end=" ")
        records = get_census_sources(conn, year)
        records_by_year[year] = records
        total_records += len(records)
        print(f"found {len(records)} sources")

    conn.close()

    if total_records == 0:
        print("No census sources found.", file=sys.stderr)
        return 1

    # Create XLSX
    print(f"Creating Excel file: {args.output}")
    create_xlsx(records_by_year, args.output)

    print(f"Done! Extracted {total_records} census sources to {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
