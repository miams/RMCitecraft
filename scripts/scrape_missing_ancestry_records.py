#!/usr/bin/env python3
"""
Scrape metadata for records that have ancestry_url in ww2_draft_updated.xlsx
but whose RIN is not yet in ww2-draft.db.

Run from project root:
    uv run python scripts/scrape_missing_ancestry_records.py
"""
import asyncio
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from loguru import logger

from rmcitecraft.database.draft_registration_db import (
    DraftRegistrationRepository,
    DRAFT_DB_PATH,
)
from rmcitecraft.services.ancestrylibrary_draft_scraper import AncestryLibraryDraftScraper


EXCEL_PATH = Path("ww2_draft_updated.xlsx")
DB_PATH = DRAFT_DB_PATH


def load_missing_records():
    """Return rows from Excel with ancestry_url but RIN absent from ww2-draft.db."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT rin FROM draft_registration WHERE rin IS NOT NULL")
    db_rins = set(r[0] for r in cursor.fetchall())
    conn.close()

    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    col = {h: headers.index(h) for h in headers}

    missing = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[col["ancestry_url"]]
        rin = row[col["rin"]]
        if url and str(url).startswith("http") and rin and int(rin) not in db_rins:
            missing.append(
                {
                    "rin": int(rin),
                    "given_name": row[col["given_name"]] or "",
                    "surname": row[col["surname"]] or "",
                    "ancestry_url": str(url).strip(),
                    "state": row[col["state"]] or "",
                }
            )

    wb.close()
    return missing, db_rins


async def scrape_all():
    """Main scraping loop — metadata only, no image downloads."""
    logger.info("Loading missing records from Excel...")
    records, _ = load_missing_records()
    logger.info(f"Found {len(records)} records to scrape")

    if not records:
        logger.info("Nothing to do — all Excel RINs already present in ww2-draft.db.")
        return

    repo = DraftRegistrationRepository(DB_PATH)
    batch_id = repo.create_batch(
        batch_name=f"Missing ancestry records {datetime.now(timezone.utc).strftime('%Y-%m-%d')}",
        notes=f"Scrape {len(records)} records absent from ww2-draft.db; metadata only",
    )
    logger.info(f"Created batch ID: {batch_id}")

    scraper = AncestryLibraryDraftScraper()
    connected = await scraper.connect()
    if not connected:
        logger.error(
            "Failed to connect to Chrome CDP.  "
            "Is Chrome running with --remote-debugging-port=9222?"
        )
        return

    success_count = 0
    fail_count = 0
    failed_records: list[tuple] = []

    try:
        for idx, rec in enumerate(records, start=1):
            rin = rec["rin"]
            url = rec["ancestry_url"]
            given = rec["given_name"]
            surname = rec["surname"]

            logger.info(f"\n[{idx}/{len(records)}] RIN {rin}: {given} {surname}")
            logger.info(f"  URL: {url}")

            try:
                registration, _ = await scraper.scrape_and_download(
                    record_url=url,
                    rin=rin,
                    metadata_only=True,
                )

                if registration is None:
                    logger.warning(f"  ⚠️  No data returned for RIN {rin} — skipping")
                    fail_count += 1
                    failed_records.append((rin, given, surname, "No data returned"))
                    continue

                # Attach RIN and mark as pre-linked from the spreadsheet
                registration.batch_id = batch_id
                registration.rin = rin
                registration.rin_link_status = "linked"
                registration.rin_link_method = "auto"
                registration.rin_link_notes = "Pre-linked from ww2_draft_updated.xlsx"
                registration.rin_linked_at = datetime.now(timezone.utc).isoformat()

                reg_id = repo.insert_registration(registration)
                logger.info(
                    f"  ✅ Saved reg_id={reg_id} — {registration.full_name}"
                )
                success_count += 1

            except Exception as exc:
                logger.error(f"  ❌ Error for RIN {rin}: {exc}", exc_info=True)
                fail_count += 1
                failed_records.append((rin, given, surname, str(exc)))

    finally:
        await scraper.disconnect()

    logger.info("\n" + "=" * 60)
    logger.info("SCRAPING COMPLETE")
    logger.info(f"  Successful : {success_count}")
    logger.info(f"  Failed     : {fail_count}")
    logger.info(f"  Total      : {len(records)}")

    if failed_records:
        logger.info("\nFailed records:")
        for rin, given, surname, reason in failed_records:
            logger.info(f"  RIN {rin:>6}: {given} {surname} — {reason}")


if __name__ == "__main__":
    asyncio.run(scrape_all())
