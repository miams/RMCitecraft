#!/usr/bin/env python3
"""
Analyze World War II Draft Registration spreadsheet file.

Provides comprehensive breakdown of the familysearch_citation and ancestry_url columns
to understand URL coverage and identify records ready for automation.

Usage:
    analyze-draft-spreadsheet                    # Uses default: ww2_draft_updated.xlsx
    analyze-draft-spreadsheet <file.xlsx>        # Analyze specific file
    analyze-draft-spreadsheet ~/.rmcitecraft/uploads/ww2_draft_updated.xlsx
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, Any, Tuple

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: uv pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.text import Text


def analyze_spreadsheet(filepath: Path) -> Dict[str, Any]:
    """
    Analyze draft registration spreadsheet.

    Returns:
        Dictionary with analysis results
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    try:
        fs_col_idx = headers.index('familysearch_citation')
        ancestry_col_idx = headers.index('ancestry_url')
    except ValueError as e:
        raise ValueError(f"Missing required column: {e}. Expected 'familysearch_citation' and 'ancestry_url'")

    # Patterns
    fs_url_pattern = re.compile(r'https?://(?:www\.)?familysearch\.org/[^\s]+|ark:/61903/[^\s,;]+', re.IGNORECASE)
    ancestry_url_pattern = re.compile(r'https?://(?:www\.)?ancestrylibrary\.com/[^\s]+', re.IGNORECASE)
    any_url_pattern = re.compile(r'https?://[^\s]+', re.IGNORECASE)

    # Counters
    total_records = 0
    fs_empty = 0
    fs_url_only = 0
    fs_url_in_citation = 0
    fs_ancestry_url = 0
    fs_text_no_url = 0

    ancestry_same_in_fs = 0
    ancestry_different_in_fs = 0
    ancestry_total = 0
    ancestry_url_with_fs_empty = 0

    # Draft registration type counters
    draft_type_young_men = 0  # /2238/
    draft_type_1942 = 0       # /1002/
    draft_type_other_rows = []  # Track rows with other/unknown types

    # Examples
    examples = {
        'fs_url_only': [],
        'fs_url_in_citation': [],
        'fs_text_no_url': [],
        'ancestry_same': [],
        'ancestry_different': []
    }

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        total_records += 1
        fs_val = row[fs_col_idx]
        ancestry_val = row[ancestry_col_idx]

        fs_text = str(fs_val).strip() if fs_val else ""
        ancestry_text = str(ancestry_val).strip() if ancestry_val else ""

        # Analyze familysearch_citation column
        if not fs_text:
            fs_empty += 1
        else:
            has_fs_url = fs_url_pattern.search(fs_text)
            has_ancestry_url = ancestry_url_pattern.search(fs_text)
            has_any_url = any_url_pattern.search(fs_text)

            if has_ancestry_url:
                fs_ancestry_url += 1
            elif has_fs_url:
                cleaned = fs_text.strip().strip('"\'')
                if fs_url_pattern.fullmatch(cleaned):
                    fs_url_only += 1
                    if len(examples['fs_url_only']) < 2:
                        examples['fs_url_only'].append((row_num, fs_text[:100]))
                else:
                    fs_url_in_citation += 1
                    if len(examples['fs_url_in_citation']) < 2:
                        examples['fs_url_in_citation'].append((row_num, fs_text[:120]))
            elif has_any_url:
                fs_text_no_url += 1
            else:
                fs_text_no_url += 1
                if len(examples['fs_text_no_url']) < 2 and len(fs_text) < 100:
                    examples['fs_text_no_url'].append((row_num, fs_text))

        # Analyze ancestry_url column
        if ancestry_text and ancestry_url_pattern.search(ancestry_text):
            ancestry_total += 1

            # Check if familysearch_citation is empty
            if not fs_text:
                ancestry_url_with_fs_empty += 1

            if fs_text and ancestry_text.lower() in fs_text.lower():
                ancestry_same_in_fs += 1
                if len(examples['ancestry_same']) < 2:
                    examples['ancestry_same'].append((row_num, ancestry_text[:80]))
            else:
                ancestry_different_in_fs += 1
                if len(examples['ancestry_different']) < 2:
                    examples['ancestry_different'].append((row_num, (fs_text[:60] if fs_text else "EMPTY", ancestry_text[:60])))

            # Analyze draft registration type
            if '/2238/' in ancestry_text:
                draft_type_young_men += 1
            elif '/1002/' in ancestry_text:
                draft_type_1942 += 1
            else:
                # Track "Other/Unknown" types
                draft_type_other_rows.append((row_num, ancestry_text[:100]))

    wb.close()

    return {
        'total_records': total_records,
        'fs_empty': fs_empty,
        'fs_url_only': fs_url_only,
        'fs_url_in_citation': fs_url_in_citation,
        'fs_ancestry_url': fs_ancestry_url,
        'fs_text_no_url': fs_text_no_url,
        'ancestry_same_in_fs': ancestry_same_in_fs,
        'ancestry_different_in_fs': ancestry_different_in_fs,
        'ancestry_total': ancestry_total,
        'ancestry_url_with_fs_empty': ancestry_url_with_fs_empty,
        'draft_type_young_men': draft_type_young_men,
        'draft_type_1942': draft_type_1942,
        'draft_type_other_rows': draft_type_other_rows,
        'examples': examples
    }


def display_results(results: Dict[str, Any], filepath: Path, console: Console):
    """Display analysis results using Rich formatting."""

    total = results['total_records']

    # Header
    console.print()
    console.print(Panel(
        f"[bold cyan]Draft Registration Spreadsheet Analysis[/bold cyan]\n"
        f"File: {filepath}\n"
        f"Total Records: {total:,}",
        border_style="cyan"
    ))

    # FamilySearch Citation breakdown
    console.print("\n[bold yellow]═══ FAMILYSEARCH_CITATION COLUMN BREAKDOWN ═══[/bold yellow]\n")

    fs_table = Table(show_header=True, header_style="bold magenta", border_style="blue")
    fs_table.add_column("Category", style="cyan", width=40)
    fs_table.add_column("Count", justify="right", style="green", width=10)
    fs_table.add_column("%", justify="right", style="yellow", width=8)

    # Calculate subtotal for URL categories
    url_subtotal = results['fs_url_in_citation'] + results['fs_url_only'] + results['fs_ancestry_url']

    fs_table.add_row("FamilySearch URL as part of citation", f"{results['fs_url_in_citation']:,}", f"{results['fs_url_in_citation']/total*100:.1f}%")
    fs_table.add_row("FamilySearch URL only", f"{results['fs_url_only']:,}", f"{results['fs_url_only']/total*100:.1f}%")
    fs_table.add_row("Ancestry URL", f"{results['fs_ancestry_url']:,}", f"{results['fs_ancestry_url']/total*100:.1f}%")
    fs_table.add_row("─" * 40, "─" * 10, "─" * 8, style="dim")
    fs_table.add_row("[bold]Subtotal (with URLs)[/bold]", f"[bold]{url_subtotal:,}[/bold]", f"[bold]{url_subtotal/total*100:.1f}%[/bold]")
    fs_table.add_row("─" * 40, "─" * 10, "─" * 8, style="dim")
    fs_table.add_row("Text but no URL", f"{results['fs_text_no_url']:,}", f"{results['fs_text_no_url']/total*100:.1f}%")
    fs_table.add_row("Empty", f"{results['fs_empty']:,}", f"{results['fs_empty']/total*100:.1f}%")
    fs_table.add_row("─" * 40, "─" * 10, "─" * 8, style="dim")
    fs_table.add_row("[bold]TOTAL[/bold]", f"[bold]{total:,}[/bold]", "[bold]100.0%[/bold]")

    console.print(fs_table)

    # Ancestry URL breakdown
    if results['ancestry_total'] > 0:
        console.print("\n[bold yellow]═══ ANCESTRY_URL COLUMN BREAKDOWN ═══[/bold yellow]\n")

        ancestry_table = Table(show_header=True, header_style="bold magenta", border_style="blue")
        ancestry_table.add_column("Category", style="cyan", width=40)
        ancestry_table.add_column("Count", justify="right", style="green", width=10)
        ancestry_table.add_column("% of URLs", justify="right", style="yellow", width=12)

        ancestry_table.add_row(
            "Same URL in familysearch_citation",
            f"{results['ancestry_same_in_fs']:,}",
            f"{results['ancestry_same_in_fs']/results['ancestry_total']*100:.1f}%"
        )
        ancestry_table.add_row(
            "Different from familysearch_citation",
            f"{results['ancestry_different_in_fs']:,}",
            f"{results['ancestry_different_in_fs']/results['ancestry_total']*100:.1f}%"
        )
        ancestry_table.add_row(
            "Ancestry_URL present with familysearch_citation empty",
            f"{results['ancestry_url_with_fs_empty']:,}",
            f"{results['ancestry_url_with_fs_empty']/results['ancestry_total']*100:.1f}%"
        )
        ancestry_table.add_row("─" * 40, "─" * 10, "─" * 12, style="dim")
        ancestry_table.add_row(
            "[bold]TOTAL with Ancestry URL[/bold]",
            f"[bold]{results['ancestry_total']:,}[/bold]",
            "[bold]100.0%[/bold]"
        )

        console.print(ancestry_table)

    # Draft Registration Type breakdown
    if results['ancestry_total'] > 0:
        console.print("\n[bold yellow]═══ DRAFT REGISTRATION TYPE ═══[/bold yellow]\n")

        draft_type_table = Table(show_header=True, header_style="bold magenta", border_style="blue")
        draft_type_table.add_column("Type", style="cyan", width=60)
        draft_type_table.add_column("Count", justify="right", style="green", width=10)
        draft_type_table.add_column("% of URLs", justify="right", style="yellow", width=12)

        draft_type_table.add_row(
            "U.S., World War II Draft Cards Young Men, 1940-1947 (/2238/)",
            f"{results['draft_type_young_men']:,}",
            f"{results['draft_type_young_men']/results['ancestry_total']*100:.1f}%" if results['ancestry_total'] > 0 else "0.0%"
        )
        draft_type_table.add_row(
            "U.S., World War II Draft Registration Cards, 1942 (/1002/)",
            f"{results['draft_type_1942']:,}",
            f"{results['draft_type_1942']/results['ancestry_total']*100:.1f}%" if results['ancestry_total'] > 0 else "0.0%"
        )

        # Calculate "other" types
        other_types = results['ancestry_total'] - results['draft_type_young_men'] - results['draft_type_1942']
        draft_type_table.add_row(
            "Other/Unknown",
            f"{other_types:,}",
            f"{other_types/results['ancestry_total']*100:.1f}%" if results['ancestry_total'] > 0 else "0.0%"
        )

        draft_type_table.add_row("─" * 60, "─" * 10, "─" * 12, style="dim")
        draft_type_table.add_row(
            "[bold]TOTAL with Ancestry URL[/bold]",
            f"[bold]{results['ancestry_total']:,}[/bold]",
            "[bold]100.0%[/bold]"
        )

        console.print(draft_type_table)

        # Display Other/Unknown rows if any
        if results['draft_type_other_rows']:
            console.print("\n[cyan]Other/Unknown Draft Registration Types:[/cyan]")
            for row_num, url in results['draft_type_other_rows'][:10]:  # Show first 10
                console.print(f"  Row {row_num}: {url}")
            if len(results['draft_type_other_rows']) > 10:
                console.print(f"  ... and {len(results['draft_type_other_rows']) - 10} more rows")

    # Summary totals
    console.print("\n[bold yellow]═══ SUMMARY TOTALS ═══[/bold yellow]\n")

    fs_with_url = results['fs_url_only'] + results['fs_url_in_citation'] + results['fs_ancestry_url']
    no_urls = results['fs_empty'] - results['ancestry_total']

    summary_table = Table(show_header=True, header_style="bold magenta", border_style="blue")
    summary_table.add_column("Metric", style="cyan", width=50)
    summary_table.add_column("Count", justify="right", style="green", width=10)

    summary_table.add_row("[bold]FamilySearch_citation with ANY URL[/bold]", f"[bold green]{fs_with_url:,}[/bold green]")
    summary_table.add_row("[bold]Ancestry_url with URL present[/bold]", f"[bold green]{results['ancestry_total']:,}[/bold green]")
    summary_table.add_row("[bold]Records with NO URLs anywhere[/bold]", f"[bold yellow]{no_urls:,}[/bold yellow]")

    console.print(summary_table)

    # Key insights
    console.print("\n[bold yellow]═══ KEY INSIGHTS ═══[/bold yellow]\n")

    insights = [
        f"• {results['fs_empty']/total*100:.1f}% of records ({results['fs_empty']:,}) have no citation - candidates for Ancestry URL discovery",
        f"• {fs_with_url:,} records ({fs_with_url/total*100:.1f}%) have some kind of URL in familysearch_citation",
        f"• {results['ancestry_total']:,} records ({results['ancestry_total']/total*100:.1f}%) have Ancestry URLs in the ancestry_url column",
        f"• {no_urls:,} records ({no_urls/total*100:.1f}%) have absolutely no URLs - pure discovery candidates",
    ]

    if results['ancestry_different_in_fs'] > 0:
        insights.append(
            f"• {results['ancestry_different_in_fs']:,} records have Ancestry URLs that differ between columns (potential data quality issue)"
        )

    for insight in insights:
        console.print(insight)

    # Examples
    console.print("\n[bold yellow]═══ EXAMPLES ═══[/bold yellow]\n")

    if results['examples']['fs_url_only']:
        console.print("[cyan]FamilySearch URL only:[/cyan]")
        for row_num, text in results['examples']['fs_url_only']:
            console.print(f"  Row {row_num}: {text}")

    if results['examples']['fs_url_in_citation']:
        console.print("\n[cyan]FamilySearch URL in citation:[/cyan]")
        for row_num, text in results['examples']['fs_url_in_citation']:
            console.print(f"  Row {row_num}: {text}...")

    if results['examples']['fs_text_no_url']:
        console.print("\n[cyan]Text but no URL:[/cyan]")
        for row_num, text in results['examples']['fs_text_no_url']:
            console.print(f"  Row {row_num}: {text}")

    console.print()


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Analyze World War II Draft Registration spreadsheet",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  analyze-draft-spreadsheet                    # Uses default: ww2_draft_updated.xlsx
  analyze-draft-spreadsheet my_file.xlsx       # Specify a different file
  analyze-draft-spreadsheet ~/.rmcitecraft/uploads/ww2_draft_updated.xlsx
        """
    )

    parser.add_argument(
        'file',
        type=str,
        nargs='?',
        default='ww2_draft_updated.xlsx',
        help='Path to Excel file (.xlsx) (default: ww2_draft_updated.xlsx)'
    )

    args = parser.parse_args()

    # Resolve file path
    filepath = Path(args.file).expanduser()

    if not filepath.exists():
        # Try ~/.rmcitecraft/uploads/ as fallback
        fallback = Path.home() / '.rmcitecraft' / 'uploads' / args.file
        if fallback.exists():
            filepath = fallback
        else:
            print(f"Error: File not found: {filepath}", file=sys.stderr)
            return 1

    if not filepath.suffix.lower() in ['.xlsx', '.xls']:
        print(f"Error: File must be an Excel file (.xlsx or .xls): {filepath}", file=sys.stderr)
        return 1

    console = Console()

    try:
        results = analyze_spreadsheet(filepath)
        display_results(results, filepath, console)
        return 0
    except Exception as e:
        console.print(f"[bold red]Error analyzing spreadsheet:[/bold red] {e}", style="red")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
